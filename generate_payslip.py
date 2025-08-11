#!/usr/bin/env python3
import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

EXCEL_FILE_PATH = "/workspace/salary.xlsx"
SHEET_NAME_OR_INDEX = 0
HEADER_ROW_INDEX_ONE_BASED = 2
OUTPUT_DIR = "./output"
CURRENCY_SYMBOL = "₹"
COMPANY_NAME = ""
MONTH_LABEL_DEFAULT = datetime.now().strftime("%b %Y")


def prompt_with_default(prompt_text: str, default_value: Optional[str]) -> str:
    suffix = f" [{default_value}]" if default_value not in (None, "") else ""
    value = input(f"{prompt_text}{suffix}: ").strip()
    return value or (default_value or "")


def normalize_column(name: str) -> str:
    return "".join(ch for ch in str(name).strip().lower() if ch.isalnum())


def find_code_column(columns: List[str]) -> Optional[str]:
    normalized_to_original: Dict[str, str] = {normalize_column(c): c for c in columns}
    candidates = [
        "code",
        "employeecode",
        "empcode",
        "employeeno",
        "employeenumber",
        "ecode",
    ]
    for cand in candidates:
        if cand in normalized_to_original:
            return normalized_to_original[cand]
    for n, original in normalized_to_original.items():
        if "code" in n:
            return original
    return None


def format_amount(value: Any, currency_symbol: str) -> str:
    if value is None or value == "":
        return "-"
    try:
        number = float(value)
    except Exception:
        return str(value)
    if number.is_integer():
        return f"{currency_symbol} {int(number):,}"
    return f"{currency_symbol} {number:,.2f}"


def read_sheet_headers_and_rows(
    excel_path: str,
    sheet_name_or_index: Any,
    header_row_one_based: Optional[int] = None,
) -> Tuple[List[str], List[List[Any]]]:
    wb = load_workbook(excel_path, data_only=True)
    if isinstance(sheet_name_or_index, int) or (isinstance(sheet_name_or_index, str) and sheet_name_or_index.isdigit()):
        idx = int(sheet_name_or_index)
        ws = wb[wb.sheetnames[idx]]
    else:
        ws = wb[str(sheet_name_or_index)]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_row: Optional[List[Any]] = None
    header_index: Optional[int] = None

    # Strategy 1: Use the provided 1-based header row index if given
    if header_row_one_based is not None and 1 <= header_row_one_based <= len(rows):
        header_index = header_row_one_based - 1
        header_row = [str(c) if c is not None else "" for c in rows[header_index]]

    def row_has_code(cells: List[Any]) -> bool:
        return any(str(v).strip().lower() == "code" for v in cells if v is not None)

    # Strategy 2: If selected header row doesn't contain 'Code', try to find a nearby row that does
    if header_row is None or not row_has_code(header_row):
        search_limit = min(10, len(rows))
        for i in range(search_limit):
            candidate = rows[i]
            if any(cell is not None and str(cell).strip() != "" for cell in candidate):
                if row_has_code(candidate):
                    header_row = [str(c) if c is not None else "" for c in candidate]
                    header_index = i
                    break

    # Strategy 3: Fallback to first non-empty row
    if header_row is None:
        for i, r in enumerate(rows):
            if any(cell is not None and str(cell).strip() != "" for cell in r):
                header_row = [str(c) if c is not None else "" for c in r]
                header_index = i
                break

    if header_row is None or header_index is None:
        return [], []

    data_rows = rows[header_index + 1 :]
    header_len = len(header_row)
    normalized_rows: List[List[Any]] = []
    for r in data_rows:
        row_values = list(r) + [None] * (header_len - len(r))
        normalized_rows.append(row_values[:header_len])

    return header_row, normalized_rows


def find_matching_row(headers: List[str], rows: List[List[Any]], code_col: str, target_code: str) -> Optional[int]:
    if code_col not in headers:
        return None
    col_idx = headers.index(code_col)
    for i, r in enumerate(rows):
        value = r[col_idx]
        if value is None:
            continue
        if str(value).strip().lower() == target_code:
            return i
    return None


def row_to_key_values(headers: List[str], row: List[Any], currency_symbol: str) -> List[Tuple[str, str]]:
    kv: List[Tuple[str, str]] = []
    for col, value in zip(headers, row):
        if isinstance(value, (int, float)) and normalize_column(col) not in {"slno", "sl", "dp", "glis"}:
            display = format_amount(value, currency_symbol)
        else:
            display = "-" if value in (None, "") else str(value)
        kv.append((str(col), display))
    return kv


def make_table_data(kv_pairs: List[Tuple[str, str]]) -> List[List[str]]:
    rows: List[List[str]] = [["Field", "Value"]]
    for key, value in kv_pairs:
        rows.append([key, value])
    return rows


def register_unicode_font() -> tuple[str, Optional[str]]:
    """Register a Unicode-capable font and return (regular_name, bold_name).

    Tries common system fonts. Falls back to built-in if none found.
    """
    candidates = [
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "DejaVuSans-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("NotoSans", "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf", "NotoSans-Bold", "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),
        ("FreeSans", "/usr/share/fonts/truetype/freefont/FreeSans.ttf", "FreeSansBold", "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf"),
    ]

    for reg_name, reg_path, bold_name, bold_path in candidates:
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(reg_name, reg_path))
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                else:
                    bold_name = None
                return reg_name, bold_name
            except Exception:
                continue

    # Fallback: built-in Helvetica (may not support ₹)
    return "Helvetica", "Helvetica-Bold"


def build_pdf(
    output_pdf_path: str,
    header_title: str,
    company_name: Optional[str],
    body_flowables: List[Any],
) -> None:
    os.makedirs(os.path.dirname(output_pdf_path), exist_ok=True)

    font_name, bold_font_name = register_unicode_font()

    doc = SimpleDocTemplate(
        output_pdf_path,
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=header_title,
        author=company_name or "",
    )

    styles = getSampleStyleSheet()
    # Apply font to common styles
    for style_key in ["Normal", "BodyText", "Title", "Heading1", "Heading2", "Heading3"]:
        if style_key in styles:
            styles[style_key].fontName = bold_font_name if (style_key.startswith("Heading") or style_key == "Title") and bold_font_name else font_name
    story: List[Any] = []

    if company_name:
        story.append(Paragraph(f"<b>{company_name}</b>", styles["Title"]))
        story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(header_title, styles["Heading2"]))
    story.append(Spacer(1, 4 * mm))

    story.extend(body_flowables)

    doc.build(story)


def main() -> None:
    print("Interactive Payslip PDF Generator")

    excel_path = EXCEL_FILE_PATH
    sheet_input = SHEET_NAME_OR_INDEX
    if isinstance(sheet_input, str) and sheet_input.isdigit():
        sheet_name: Any = int(sheet_input)
    else:
        sheet_name = sheet_input
    employee_code = prompt_with_default("Employee code", "").strip()
    if not employee_code:
        raise SystemExit("Employee code is required.")
    company_name = COMPANY_NAME
    month = MONTH_LABEL_DEFAULT
    output_dir = OUTPUT_DIR
    currency_symbol = CURRENCY_SYMBOL

    headers, rows = read_sheet_headers_and_rows(
        excel_path, sheet_name, header_row_one_based=HEADER_ROW_INDEX_ONE_BASED
    )
    if not headers or not rows:
        raise SystemExit("The selected sheet appears to be empty or has no headers.")

    code_col = "Code"
    if code_col not in headers:
        raise SystemExit("The sheet must contain a 'Code' column.")

    target_code = employee_code.strip().lower()
    match_idx = find_matching_row(headers, rows, code_col, target_code)
    if match_idx is None:
        sample_values = ", ".join(sorted({str(r[headers.index(code_col)]) for r in rows if r[headers.index(code_col)] is not None})[:20])
        raise SystemExit(
            f"No employee found with code '{employee_code}'. Available samples: {sample_values} ..."
        )

    row = rows[match_idx]

    # Classify columns into income and deduction sections
    norm_to_col = {normalize_column(h): h for h in headers}
    deduction_keys = {
        "pf",
        "incometax",
        "glis",
        "refundspf",
        "refundswf",
        "tds",
        "esi",
        "professionaltax",
        "loan",
        "advance",
        "totaldeductions",
    }
    summary_gross_keys = {"gross"}
    summary_total_ded_keys = {"totaldeductions"}
    summary_net_keys = {"netpay"}
    info_keys = {"slno", "sl", "code", "name", "employee", "employeename", "empname"}

    incomes: List[Tuple[str, Any]] = []
    deductions: List[Tuple[str, Any]] = []
    gross_value = None
    total_deductions_value = None
    net_pay_value = None

    for h in headers:
        n = normalize_column(h)
        value = row[headers.index(h)]
        if n in info_keys:
            continue
        if n in summary_net_keys:
            net_pay_value = value
            continue
        if n in summary_gross_keys:
            gross_value = value
            continue
        if n in summary_total_ded_keys:
            total_deductions_value = value
            continue
        # Classify unknown numeric columns: assume known deductions list, else income
        if n in deduction_keys:
            deductions.append((h, value))
        else:
            incomes.append((h, value))

    # Compute totals if not provided
    if gross_value is None:
        gross_value = sum((v or 0) for _, v in incomes if isinstance(v, (int, float)))
    if total_deductions_value is None:
        total_deductions_value = sum((v or 0) for _, v in deductions if isinstance(v, (int, float)))
    if net_pay_value is None:
        try:
            net_pay_value = float(gross_value or 0) - float(total_deductions_value or 0)
        except Exception:
            net_pay_value = None

    # Build income and deduction tables
    def build_section(title: str, items: List[Tuple[str, Any]], total_label: str, total_value: Any) -> Table:
        rows_data: List[List[str]] = [[f"{title}", "Amount"]]
        for label, val in items:
            rows_data.append([str(label), format_amount(val, currency_symbol)])
        rows_data.append([total_label, format_amount(total_value, currency_symbol)])

        table = Table(rows_data, colWidths=[80 * mm, None])
        font_name, bold_font_name = register_unicode_font()
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTNAME", (0, 0), (-1, 0), bold_font_name or font_name),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ])
        # Stripe body rows
        for idx in range(1, len(rows_data)):
            if idx % 2 == 0:
                style.add("BACKGROUND", (0, idx), (-1, idx), colors.whitesmoke)
        # Bold total row
        style.add("FONTNAME", (0, len(rows_data) - 1), (-1, len(rows_data) - 1), bold_font_name or font_name)
        table.setStyle(style)
        return table

    income_table = build_section("Income", incomes, "Gross", gross_value)
    deduction_table = build_section("Deductions", deductions, "Total Deductions", total_deductions_value)

    # Place the two sections horizontally
    sections = Table([[income_table, deduction_table]], colWidths=[100 * mm, None])
    sections.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    # Net Pay paragraph outside tables
    font_name, bold_font_name = register_unicode_font()
    styles = getSampleStyleSheet()
    styles["Heading3"].fontName = bold_font_name or font_name
    net_pay_para = Paragraph(f"<b>Net Pay:</b> {format_amount(net_pay_value, currency_symbol)}", styles["Heading3"])

    body_flowables = [sections, Spacer(1, 6 * mm), net_pay_para]

    emp_identifier = str(row[headers.index(code_col)])
    emp_name = None
    for c in headers:
        if normalize_column(c) in {"name", "employeename", "empname"}:
            emp_name = str(row[headers.index(c)])
            break

    header_title = (
        f"Salary Statement for {month}"
        + (f" — {emp_name}" if emp_name else "")
        + f" (Code: {emp_identifier})"
    )

    safe_name = (emp_name or "").strip().replace(" ", "_") if emp_name else "unknown"
    output_filename = f"payslip_{safe_name}_{emp_identifier}.pdf"
    output_path = os.path.join(output_dir, output_filename)

    build_pdf(
        output_pdf_path=output_path,
        header_title=header_title,
        company_name=company_name,
        body_flowables=body_flowables,
    )

    print(f"Generated: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
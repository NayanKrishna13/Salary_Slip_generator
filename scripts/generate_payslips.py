#!/usr/bin/env python3
import argparse
import math
import os
import csv
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Mapping

from PIL import Image, ImageDraw, ImageFont
from num2words import num2words
from openpyxl import load_workbook


DEFAULT_FONT_REGULAR_CANDIDATES = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
]
DEFAULT_FONT_BOLD_CANDIDATES = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
]


def load_font(preferred_path: Optional[str], fallback_candidates: List[str], size: int) -> ImageFont.FreeTypeFont:
    if preferred_path and os.path.isfile(preferred_path):
        try:
            return ImageFont.truetype(preferred_path, size=size)
        except Exception:
            pass
    for path in fallback_candidates:
        if os.path.isfile(path):
            try:
                return ImageFont.truetype(path, size=size)
            except Exception:
                continue
    # Final fallback
    return ImageFont.load_default()


def format_currency(value: float) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        value = 0.0
    return f"{float(value):,.2f}"


def amount_in_words(value: float) -> str:
    try:
        integer = int(round(float(value)))
        words = num2words(integer, lang="en_IN").replace(" and ", " ")
        words = words.replace(",", "")
        words = words.title()
        return f"Rupees {words} Only"
    except Exception:
        return "Rupees Zero Only"


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def safe_get(row: Mapping[str, Any], key: str, default: str = "") -> str:
    if key in row and not _is_missing(row[key]):
        return str(row[key])
    return default


def safe_get_any(row: Mapping[str, Any], keys: List[str], default: str = "") -> str:
    for key in keys:
        if key in row and not _is_missing(row[key]):
            return str(row[key])
    return default


def safe_get_num(row: Mapping[str, Any], key: str, default: float = 0.0) -> float:
    if key in row and not _is_missing(row[key]):
        try:
            value = row[key]
            if isinstance(value, str):
                value = value.replace(",", "").strip()
            return float(value)
        except Exception:
            return default
    return default


def draw_centered_text(draw: ImageDraw.ImageDraw, text: str, center_x: int, y: int, font: ImageFont.FreeTypeFont, fill: Tuple[int, int, int] = (0, 0, 0)) -> int:
    bbox = draw.textbbox((0, 0), text, font=font)
    width = bbox[2] - bbox[0]
    draw.text((center_x - width // 2, y), text, font=font, fill=fill)
    return bbox[3] - bbox[1]


def draw_key_value(draw: ImageDraw.ImageDraw, x_key: int, x_val: int, y: int, key_text: str, val_text: str, key_font: ImageFont.FreeTypeFont, val_font: ImageFont.FreeTypeFont, key_fill=(0, 0, 0), val_fill=(0, 0, 0)) -> int:
    draw.text((x_key, y), key_text, font=key_font, fill=key_fill)
    draw.text((x_val, y), val_text, font=val_font, fill=val_fill)
    bbox = draw.textbbox((x_val, y), val_text, font=val_font)
    return bbox[3] - bbox[1]


def read_table(input_path: str, use_first_row_as_header: bool = True) -> Tuple[List[Dict[str, Any]], List[str]]:
    ext = os.path.splitext(input_path.lower())[1]

    rows_raw: List[List[Any]] = []
    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(filename=input_path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            rows_raw.append(list(row))
    elif ext == ".xls":
        raise ValueError(".xls format is not supported without pandas; please convert to .xlsx")
    elif ext in [".csv", ".tsv"]:
        delimiter = "," if ext == ".csv" else "\t"
        with open(input_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows_raw.append(row)
    else:
        raise ValueError(f"Unsupported input file type: {ext}")

    # Find first non-empty row to use as header (if requested)
    header_idx = 0
    if use_first_row_as_header:
        found = False
        for i, raw in enumerate(rows_raw):
            if any(not _is_missing(cell if not isinstance(cell, str) else cell.strip()) for cell in raw):
                header_idx = i
                found = True
                break
        if not found:
            return [], []
    else:
        header_idx = 0

    # Build header names
    raw_header = rows_raw[header_idx] if rows_raw else []
    header: List[str] = []
    for j, name in enumerate(raw_header):
        text = str(name).strip() if not _is_missing(name) else f"Column{j+1}"
        header.append(text)

    # Build list of row dicts after header row
    result_rows: List[Dict[str, Any]] = []
    for raw in rows_raw[header_idx + 1:]:
        # Skip entirely empty rows
        if all(_is_missing(cell if not isinstance(cell, str) else cell.strip()) for cell in raw):
            continue
        row_dict: Dict[str, Any] = {}
        for j, key in enumerate(header):
            value = raw[j] if j < len(raw) else None
            row_dict[key] = value
        result_rows.append(row_dict)

    return result_rows, header


def ensure_output_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def render_payslip(
    row: Mapping[str, Any],
    output_path: str,
    logo_path: Optional[str],
    regular_font_path: Optional[str],
    bold_font_path: Optional[str],
    month_text: Optional[str] = None,
    image_width: int = 1600,
    image_height: int = 1120,
    background_color: Tuple[int, int, int] = (255, 255, 255),
    output_format: str = "png",
    output_filename: Optional[str] = None,
) -> None:
    image = Image.new("RGB", (image_width, image_height), color=background_color)
    draw = ImageDraw.Draw(image)

    # Fonts
    font_title = load_font(bold_font_path, DEFAULT_FONT_BOLD_CANDIDATES, 44)
    font_subtitle = load_font(regular_font_path, DEFAULT_FONT_REGULAR_CANDIDATES, 24)
    font_section = load_font(bold_font_path, DEFAULT_FONT_BOLD_CANDIDATES, 26)
    font_label = load_font(bold_font_path, DEFAULT_FONT_BOLD_CANDIDATES, 22)
    font_value = load_font(regular_font_path, DEFAULT_FONT_REGULAR_CANDIDATES, 22)
    font_table_header = load_font(bold_font_path, DEFAULT_FONT_BOLD_CANDIDATES, 22)
    font_table_cell = load_font(regular_font_path, DEFAULT_FONT_REGULAR_CANDIDATES, 22)
    font_footer = load_font(regular_font_path, DEFAULT_FONT_REGULAR_CANDIDATES, 18)

    margin = 40
    center_x = image_width // 2
    y = margin

    # Logo
    if logo_path and os.path.isfile(logo_path):
        try:
            logo = Image.open(logo_path).convert("RGBA")
            max_logo_h = 120
            ratio = max_logo_h / logo.height
            new_w = int(logo.width * ratio)
            logo = logo.resize((new_w, max_logo_h), Image.LANCZOS)
            image.paste(logo, (margin, y), mask=logo)
        except Exception:
            pass

    # Header
    draw_centered_text(draw, "Centre for the Study of Developing Societies", center_x, y + 10, font_title)
    y += 70
    draw_centered_text(draw, "29 Rajpur Road civil lines delhi 110054.", center_x, y, font_subtitle)
    y += 35

    # Month
    if not month_text:
        month_text = safe_get(row, "Month", "")
        if not month_text:
            # Try to infer from data if available, otherwise use current month
            now = datetime.now()
            month_text = now.strftime("%B %Y")
    y += draw_centered_text(draw, f"Payslip for the month of {month_text}", center_x, y + 10, font_section)
    y += 20

    # Details box
    box_top = y + 10
    box_left = margin
    box_right = image_width - margin
    box_bottom = box_top + 210
    draw.rectangle([box_left, box_top, box_right, box_bottom], outline=(0, 0, 0), width=2)

    # Vertical divider
    mid_x = (box_left + box_right) // 2
    draw.line([(mid_x, box_top), (mid_x, box_bottom)], fill=(0, 0, 0), width=2)

    # Left column fields
    left_fields = [
        ("Name:", safe_get(row, "Name")),
        ("Joining Date:", safe_get(row, "Joining Date")),
        ("Designation:", safe_get(row, "Designation")),
        ("Department:", safe_get(row, "Department")),
        ("Location:", safe_get(row, "Location")),
        ("Effective Work Days:", safe_get(row, "Effective Work Days")),
        ("LOP:", safe_get(row, "LOP")),
    ]

    employee_no_val = safe_get_any(row, ["Employee No", "Employee Number", "Code"]) or ""

    right_fields = [
        ("Employee No:", employee_no_val),
        ("Bank Name:", safe_get(row, "Bank Name")),
        ("Bank Account No:", safe_get(row, "Bank Account No")),
        ("PAN Number:", safe_get(row, "PAN Number")),
        ("PF No:", safe_get(row, "PF No")),
        ("PF UAN:", safe_get(row, "PF UAN")),
    ]

    row_height = 28
    pad_x = 14
    key_x_left = box_left + pad_x
    val_x_left = box_left + 220
    key_x_right = mid_x + pad_x
    val_x_right = mid_x + 220

    cursor_y = box_top + 12
    for key, val in left_fields:
        draw_key_value(draw, key_x_left, val_x_left, cursor_y, key, val, font_label, font_value)
        cursor_y += row_height

    cursor_y = box_top + 12
    for key, val in right_fields:
        draw_key_value(draw, key_x_right, val_x_right, cursor_y, key, val, font_label, font_value)
        cursor_y += row_height

    y = box_bottom + 20

    # Earnings and Deductions tables
    table_left = margin
    table_right = image_width - margin
    table_top = y
    table_bottom = table_top + 360

    draw.rectangle([table_left, table_top, table_right, table_bottom], outline=(0, 0, 0), width=2)

    # Vertical grid: Earnings on left 60%, Deductions on right 40%
    earn_right = table_left + int(0.58 * (table_right - table_left))
    draw.line([(earn_right, table_top), (earn_right, table_bottom)], fill=(0, 0, 0), width=2)

    # Column headers
    # Earnings sub-table columns: Title | Master | Actual
    earn_col1 = table_left + 10
    earn_col2 = earn_right - 200
    earn_col3 = earn_right - 60

    ded_col1 = earn_right + 10
    ded_col2 = table_right - 100

    header_height = 34

    # Earnings header row background (optional):
    draw.text((earn_col1, table_top + 6), "Earnings", font=font_table_header, fill=(0, 0, 0))
    draw.text((earn_col2 - 30, table_top + 6), "Master", font=font_table_header, fill=(0, 0, 0))
    draw.text((earn_col3 - 30, table_top + 6), "Actual", font=font_table_header, fill=(0, 0, 0))

    draw.text((ded_col1, table_top + 6), "Deductions", font=font_table_header, fill=(0, 0, 0))
    draw.text((ded_col2 - 20, table_top + 6), "Actual", font=font_table_header, fill=(0, 0, 0))

    # Horizontal divider under headers
    draw.line([(table_left, table_top + header_height), (table_right, table_top + header_height)], fill=(0, 0, 0), width=2)

    # Earnings rows
    earnings_items = [
        ("BASIC", safe_get_num(row, "BASIC_master"), safe_get_num(row, "BASIC_actual")),
        ("DA", safe_get_num(row, "DA_master"), safe_get_num(row, "DA_actual")),
        ("HRA", safe_get_num(row, "HRA_master"), safe_get_num(row, "HRA_actual")),
        ("TRANSPORT ALLOWANCE", safe_get_num(row, "TRANSPORT_ALLOWANCE_master"), safe_get_num(row, "TRANSPORT_ALLOWANCE_actual")),
        ("DA TPT", safe_get_num(row, "DA_TPT_master"), safe_get_num(row, "DA_TPT_actual")),
    ]

    deductions_items = [
        ("PF", safe_get_num(row, "PF_actual")),
        ("GLIS", safe_get_num(row, "GLIS_actual")),
        ("REFUND SPF", safe_get_num(row, "REFUND_SPF_actual")),
        ("REFUND SWF", safe_get_num(row, "REFUND_SWF_actual")),
    ]

    # Filter out zero rows (if entirely zero and empty)
    earnings_items = [item for item in earnings_items if (item[1] != 0 or item[2] != 0)]
    deductions_items = [item for item in deductions_items if item[1] != 0]

    row_y = table_top + header_height + 8
    line_height = 28

    total_earn_master = 0.0
    total_earn_actual = 0.0
    for label, master, actual in earnings_items:
        draw.text((earn_col1, row_y), label, font=font_table_cell, fill=(0, 0, 0))
        draw.text((earn_col2, row_y), format_currency(master), font=font_table_cell, fill=(0, 0, 0))
        draw.text((earn_col3, row_y), format_currency(actual), font=font_table_cell, fill=(0, 0, 0))
        total_earn_master += master
        total_earn_actual += actual
        row_y += line_height

    total_deduct = 0.0
    row_y_ded = table_top + header_height + 8
    for label, actual in deductions_items:
        draw.text((ded_col1, row_y_ded), label, font=font_table_cell, fill=(0, 0, 0))
        draw.text((ded_col2, row_y_ded), format_currency(actual), font=font_table_cell, fill=(0, 0, 0))
        total_deduct += actual
        row_y_ded += line_height

    # Totals row lines
    totals_y = table_bottom - 40
    draw.line([(table_left, totals_y), (table_right, totals_y)], fill=(0, 0, 0), width=2)

    draw.text((earn_col1, totals_y + 8), "Total Earnings:INR.", font=font_table_header, fill=(0, 0, 0))
    draw.text((earn_col2, totals_y + 8), format_currency(total_earn_master), font=font_table_header, fill=(0, 0, 0))
    draw.text((earn_col3, totals_y + 8), format_currency(total_earn_actual), font=font_table_header, fill=(0, 0, 0))

    draw.text((ded_col1, totals_y + 8), "Total Deductions:INR.", font=font_table_header, fill=(0, 0, 0))
    draw.text((ded_col2, totals_y + 8), format_currency(total_deduct), font=font_table_header, fill=(0, 0, 0))

    # Net Pay
    y = table_bottom + 20
    net_pay = total_earn_actual - total_deduct
    label_text = "Net Pay for the month:"
    draw.text((margin, y), label_text, font=font_label, fill=(0, 0, 0))
    # Bold value
    value_text = format_currency(net_pay)
    bbox_label = draw.textbbox((margin, y), label_text, font=font_label)
    draw.text((bbox_label[2] + 20, y), value_text, font=font_section, fill=(0, 0, 0))
    y += 40

    # Amount in words
    draw.text((margin, y), f"({amount_in_words(net_pay)})", font=font_value, fill=(0, 0, 0))

    # Footer
    footer_text = "This is a system generated payslip and does not require a signature"
    draw.text((center_x - draw.textlength(footer_text, font=font_footer) // 2, image_height - 50), footer_text, font=font_footer, fill=(100, 100, 100))

    # Output filename: Name and Code/Employee No
    code_val = employee_no_val
    name_val = safe_get(row, "Name", "Employee").strip()
    base = f"{name_val}_{code_val}" if code_val else name_val
    safe_base = "".join(c for c in base if c.isalnum() or c in ("-", "_"))
    if not safe_base:
        safe_base = "payslip"
    suffix = f"_{month_text.replace(' ', '_')}" if month_text else ""
    auto_filename = f"{safe_base}{suffix}"

    # Decide final path and save
    if output_filename:
        final_path = output_filename
    else:
        ext = ".pdf" if output_format.lower() == "pdf" else ".png"
        final_path = os.path.join(output_path, f"{auto_filename}{ext}")

    if output_format.lower() == "pdf":
        image.convert("RGB").save(final_path, "PDF", resolution=300.0)
    else:
        image.save(final_path)


def main():
    parser = argparse.ArgumentParser(description="Generate payslip images from an Excel/CSV sheet.")
    parser.add_argument("--excel", required=True, help="Absolute path to Excel/CSV file containing payslip data")
    parser.add_argument("--output-dir", required=True, help="Absolute path to output directory for generated images")
    parser.add_argument("--logo", required=False, default=None, help="Absolute path to logo image to paste in header")
    parser.add_argument("--font-regular", required=False, default=None, help="Optional path to a TTF font for regular text")
    parser.add_argument("--font-bold", required=False, default=None, help="Optional path to a TTF font for bold text")
    parser.add_argument("--month", required=False, default=None, help="Month label, e.g. 'July 2025'. Overrides data column if provided")
    parser.add_argument("--width", type=int, default=1600, help="Image width in pixels")
    parser.add_argument("--height", type=int, default=1120, help="Image height in pixels")
    parser.add_argument("--output-format", choices=["png", "pdf"], default="png", help="Output format for the generated image (png or pdf)")
    parser.add_argument("--output-filename", help="Optional: Specify a custom output filename (e.g., 'payslip_2023_07.pdf')")

    args = parser.parse_args()

    input_path = args.excel
    if not os.path.isabs(input_path):
        raise SystemExit("--excel must be an absolute path")
    if args.logo and not os.path.isabs(args.logo):
        raise SystemExit("--logo must be an absolute path if provided")
    if not os.path.isabs(args.output_dir):
        raise SystemExit("--output-dir must be an absolute path")

    ensure_output_dir(args.output_dir)

    rows, columns = read_table(input_path, use_first_row_as_header=True)

    required_columns = [
        "Name",
        "Joining Date",
        "Designation",
        "Department",
        "Location",
        "Effective Work Days",
        "LOP",
        # Right side may supply Code instead of Employee No
        # "Employee No",
        "Bank Name",
        "Bank Account No",
        "PAN Number",
        "PF No",
        "PF UAN",
    ]

    missing = [c for c in required_columns if c not in columns]
    if missing:
        print("Warning: Missing expected columns:", missing)

    for idx, row in enumerate(rows):
        render_payslip(
            row=row,
            output_path=args.output_dir,
            logo_path=args.logo,
            regular_font_path=args.font_regular,
            bold_font_path=args.font_bold,
            month_text=args.month,
            image_width=args.width,
            image_height=args.height,
            output_format=args.output_format,
            output_filename=args.output_filename,
        )
        print(f"Rendered payslip for row {idx}")


if __name__ == "__main__":
    main()